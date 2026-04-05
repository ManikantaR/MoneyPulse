"""
Synthetic bank statement generator for MoneyPulse.
Generates realistic CSV and XLSX files in each supported parser format.
Dates: Jan 1 – Apr 5, 2026  |  ~80-90 transactions per file
"""

import csv
import os
import random
from datetime import date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

# ── Config ────────────────────────────────────────────────────────────────────

random.seed(42)  # reproducible

START_DATE = date(2026, 1, 1)
END_DATE   = date(2026, 4, 5)

OUT_DIR        = Path(__file__).parent
WATCH_BASE     = Path.home() / "moneypulse-data" / "watch-folder"

# Slug → (csv_filename, format_label)
ACCOUNTS = {
    "chase-checking-1234": ("chase-checking-2026.csv",  "Chase Checking"),
    "chase-cc-5678":        ("chase-cc-2026.csv",         "Chase Credit Card"),
    "boa-checking-9012":    ("boa-checking-2026.csv",     "Bank of America"),
    "amex-platinum-3456":   ("amex-2026.csv",             "American Express"),
    "citi-card-7890":       ("citi-2026.csv",             "Citi"),
    "generic-savings-2468": ("generic-savings-2026.xlsx", "Generic (Excel)"),
}

# ── Merchant Data ─────────────────────────────────────────────────────────────

MERCHANTS = {
    "Food & Dining": [
        ("CHIPOTLE ONLINE 4521",        (8,   16)),
        ("STARBUCKS STORE 07821",       (5,   12)),
        ("MCDONALD'S F32541",           (4,    9)),
        ("UBER EATS",                   (12,  55)),
        ("DOORDASH*PIZZA HUT",          (15,  48)),
        ("WHOLE FOODS MARKET #0892",    (45, 145)),
        ("TRADER JOE'S #134",           (30,  90)),
        ("KROGER FUEL #0032",           (35, 110)),
        ("PANERA BREAD 203541",         (8,   18)),
        ("CHICK-FIL-A #01234",          (7,   14)),
        ("FIVE GUYS BURGERS 9921",      (10,  22)),
        ("SWEETGREEN DUPONT",           (12,  18)),
        ("SUBWAY 00412",                (6,   12)),
        ("GRUBHUB*THAI KITCHEN",        (18,  45)),
        ("DUNKIN' #301234",             (3,    8)),
    ],
    "Shopping": [
        ("AMAZON.COM*MQ4ZD2",           (12, 220)),
        ("TARGET 00012345",             (25, 180)),
        ("WALMART SUPERCENTER #2341",   (30, 150)),
        ("COSTCO WHSE #0412",           (65, 280)),
        ("BEST BUY 00012345",           (40, 450)),
        ("TJ MAXX #0345",               (18,  95)),
        ("HOMEGOODS #0129",             (25,  85)),
        ("APPLE.COM/BILL",              (1,   30)),
        ("CHEWY.COM",                   (30, 120)),
        ("ETSY.COM PURCHASE",           (12,  75)),
    ],
    "Transportation": [
        ("UBER *TRIP",                  (8,   35)),
        ("LYFT *RIDE",                  (7,   30)),
        ("SHELL OIL 57442",             (40,  75)),
        ("BP#9332114",                  (38,  72)),
        ("CHEVRON 00412",               (42,  78)),
        ("EXXONMOBIL 97123",            (36,  70)),
        ("TESLA SUPERCHARGER",          (12,  30)),
        ("METRO TRANSIT PASS",          (40,  40)),
        ("PARKING 555 MAIN ST",         (5,   25)),
        ("EZ PASS REPLENISHMENT",       (25,  50)),
    ],
    "Entertainment": [
        ("NETFLIX.COM",                 (16,  16)),
        ("SPOTIFY USA",                 (11,  11)),
        ("HULU",                        (18,  18)),
        ("DISNEY PLUS",                 (14,  14)),
        ("AMC THEATERS #0342",          (12,  55)),
        ("STEAM PURCHASE",              (5,   60)),
        ("HBO MAX",                     (16,  16)),
        ("APPLE TV+",                   (10,  10)),
        ("TICKETMASTER*CONCERT",        (45, 280)),
        ("XBOX GAME PASS",              (15,  15)),
    ],
    "Utilities & Bills": [
        ("AT&T *BILL PAYMENT",          (70, 120)),
        ("VERIZON WIRELESS",            (80, 140)),
        ("XFINITY INTERNET",            (60,  90)),
        ("DUKE ENERGY",                 (55, 180)),
        ("CITY WATER BILL",             (25,  65)),
        ("GEICO AUTO INS",              (90, 180)),
        ("STATE FARM HOME",             (85, 200)),
        ("GOOGLE ONE STORAGE",          (3,   10)),
        ("AMAZON PRIME",                (15,  15)),
        ("ADOBE CREATIVE CLOUD",        (55,  55)),
    ],
    "Healthcare": [
        ("CVS PHARMACY #04122",         (8,   95)),
        ("WALGREENS #09234",            (10,  80)),
        ("DENTAL CARE OF ARLINGTON",    (95, 450)),
        ("VISION CENTER",               (60, 300)),
        ("URGENT CARE CLINIC",          (85, 250)),
        ("PLANET FITNESS",              (10,  25)),
        ("PELOTON MEMBERSHIP",          (44,  44)),
        ("CAREFIRST BCBS PREMIUM",      (180, 480)),
    ],
    "Travel": [
        ("DELTA AIR LINES",             (120, 650)),
        ("SOUTHWEST AIRLINES",          (80, 450)),
        ("MARRIOTT HOTELS",             (95, 380)),
        ("AIRBNB * STAY",               (75, 650)),
        ("HERTZ RENTAL CAR",            (55, 220)),
        ("UNITED AIRLINES",             (140, 780)),
    ],
}

FLAT_MERCHANTS = [(m, r, cat) for cat, items in MERCHANTS.items() for m, r in items]


# ── Helpers ───────────────────────────────────────────────────────────────────

def rand_date() -> date:
    delta = (END_DATE - START_DATE).days
    return START_DATE + timedelta(days=random.randint(0, delta))

def rand_amount(lo: float, hi: float) -> float:
    return round(random.uniform(lo, hi), 2)

def fmt_date(d: date) -> str:
    return d.strftime("%m/%d/%Y")

def rand_ref() -> str:
    return str(random.randint(1000000000, 9999999999))

def build_transactions(n_expenses=75, payroll_amount=3800.00):
    """
    Returns a list of dicts with keys:
      date (date), post_date (date), merchant (str), category (str),
      amount (float, always positive), is_credit (bool)
    Sorted newest-first (realistic bank statement order).
    """
    rows = []

    # Monthly payrolls
    for month in [1, 2, 3, 4]:
        for day in [1, 16]:
            d = date(2026, month, day)
            if d <= END_DATE:
                rows.append({
                    "date": d,
                    "post_date": d,
                    "merchant": "PAYROLL DIRECT DEP ACME INC",
                    "category": "Income",
                    "amount": payroll_amount,
                    "is_credit": True,
                })

    # Monthly credit-card payment (checking accounts)
    for month in [1, 2, 3]:
        d = date(2026, month, 22)
        rows.append({
            "date": d,
            "post_date": d + timedelta(days=1),
            "merchant": "ONLINE PAYMENT THANK YOU",
            "category": "Payment",
            "amount": round(random.uniform(800, 2200), 2),
            "is_credit": True,
        })

    # Random expenses
    for _ in range(n_expenses):
        merchant, (lo, hi), cat = random.choice(FLAT_MERCHANTS)
        d = rand_date()
        rows.append({
            "date": d,
            "post_date": d + timedelta(days=random.randint(0, 2)),
            "merchant": merchant,
            "category": cat,
            "amount": rand_amount(lo, hi),
            "is_credit": False,
        })

    rows.sort(key=lambda r: r["date"], reverse=True)
    return rows


def build_cc_transactions(n_charges=75):
    """Credit-card variant — no payroll, has monthly payment."""
    rows = []

    # Monthly payment
    for month in [1, 2, 3]:
        d = date(2026, month, 5)
        rows.append({
            "date": d,
            "post_date": d + timedelta(days=1),
            "merchant": "PAYMENT THANK YOU",
            "category": "Payment",
            "amount": round(random.uniform(600, 1800), 2),
            "is_credit": True,
        })

    # Random charges
    for _ in range(n_charges):
        merchant, (lo, hi), cat = random.choice(FLAT_MERCHANTS)
        d = rand_date()
        rows.append({
            "date": d,
            "post_date": d + timedelta(days=random.randint(1, 2)),
            "merchant": merchant,
            "category": cat,
            "amount": rand_amount(lo, hi),
            "is_credit": False,
        })

    rows.sort(key=lambda r: r["date"], reverse=True)
    return rows


# ── Format Writers ────────────────────────────────────────────────────────────

def write_chase_checking(path: Path):
    """
    Transaction Date,Posting Date,Description,Category,Debit,Credit,Balance
    Debit & Credit are unsigned; one populated per row.
    """
    rows = build_transactions(n_expenses=80)
    balance = 8500.00

    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Transaction Date", "Posting Date", "Description", "Category", "Debit", "Credit", "Balance"])
        for r in rows:
            if r["is_credit"]:
                balance += r["amount"]
                w.writerow([fmt_date(r["date"]), fmt_date(r["post_date"]),
                            r["merchant"], r["category"], "", f"{r['amount']:.2f}", f"{balance:.2f}"])
            else:
                balance -= r["amount"]
                w.writerow([fmt_date(r["date"]), fmt_date(r["post_date"]),
                            r["merchant"], r["category"], f"{r['amount']:.2f}", "", f"{balance:.2f}"])


def write_chase_cc(path: Path):
    """
    Transaction Date,Post Date,Description,Category,Type,Amount
    Negative = charge, positive = payment.
    """
    rows = build_cc_transactions(n_charges=80)
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Transaction Date", "Post Date", "Description", "Category", "Type", "Amount"])
        for r in rows:
            if r["is_credit"]:
                w.writerow([fmt_date(r["date"]), fmt_date(r["post_date"]),
                            r["merchant"], r["category"], "Payment", f"{r['amount']:.2f}"])
            else:
                w.writerow([fmt_date(r["date"]), fmt_date(r["post_date"]),
                            r["merchant"], r["category"], "Sale", f"-{r['amount']:.2f}"])


def write_boa(path: Path):
    """
    Date,Reference Number,Description,Amount,Running Bal.
    Negative = debit, positive = credit.
    """
    rows = build_transactions(n_expenses=78)
    balance = 6200.00
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Date", "Reference Number", "Description", "Amount", "Running Bal."])
        for r in rows:
            ref = rand_ref()
            if r["is_credit"]:
                balance += r["amount"]
                w.writerow([fmt_date(r["date"]), ref, r["merchant"],
                            f"{r['amount']:.2f}", f"{balance:.2f}"])
            else:
                balance -= r["amount"]
                w.writerow([fmt_date(r["date"]), ref, r["merchant"],
                            f"-{r['amount']:.2f}", f"{balance:.2f}"])


def write_amex(path: Path):
    """
    Date,Description,Amount
    POSITIVE = charge (Amex convention), negative = payment/refund.
    """
    rows = build_cc_transactions(n_charges=78)
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Date", "Description", "Amount"])
        for r in rows:
            if r["is_credit"]:
                # payment = negative for Amex
                w.writerow([fmt_date(r["date"]), r["merchant"], f"-{r['amount']:.2f}"])
            else:
                w.writerow([fmt_date(r["date"]), r["merchant"], f"{r['amount']:.2f}"])


def write_citi(path: Path):
    """
    Status,Date,Description,Debit,Credit
    Unsigned separate columns.
    """
    rows = build_cc_transactions(n_charges=78)
    statuses = ["Cleared"] * 9 + ["Pending"]
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Status", "Date", "Description", "Debit", "Credit"])
        for r in rows:
            status = random.choice(statuses)
            if r["is_credit"]:
                w.writerow([status, fmt_date(r["date"]), r["merchant"], "", f"{r['amount']:.2f}"])
            else:
                w.writerow([status, fmt_date(r["date"]), r["merchant"], f"{r['amount']:.2f}", ""])


def write_generic_excel(path: Path):
    """
    Excel with Chase Checking columns on Sheet1 — parsed by ExcelParser → ChaseCheckingParser.
    """
    rows = build_transactions(n_expenses=72)
    balance = 12000.00

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ["Transaction Date", "Posting Date", "Description", "Category", "Debit", "Credit", "Balance"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in rows:
        if r["is_credit"]:
            balance += r["amount"]
            ws.append([
                fmt_date(r["date"]), fmt_date(r["post_date"]),
                r["merchant"], r["category"],
                None, r["amount"], round(balance, 2),
            ])
        else:
            balance -= r["amount"]
            ws.append([
                fmt_date(r["date"]), fmt_date(r["post_date"]),
                r["merchant"], r["category"],
                r["amount"], None, round(balance, 2),
            ])

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    wb.save(path)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    writers = {
        "chase-checking-1234": (write_chase_checking, "chase-checking-2026.csv"),
        "chase-cc-5678":        (write_chase_cc,       "chase-cc-2026.csv"),
        "boa-checking-9012":    (write_boa,             "boa-checking-2026.csv"),
        "amex-platinum-3456":   (write_amex,            "amex-2026.csv"),
        "citi-card-7890":       (write_citi,            "citi-2026.csv"),
        "generic-savings-2468": (write_generic_excel,   "generic-savings-2026.xlsx"),
    }

    print(f"Generating synthetic bank statements → {OUT_DIR}\n")

    for slug, (writer_fn, filename) in writers.items():
        out_path = OUT_DIR / filename
        writer_fn(out_path)

        # Count rows (minus header for CSV)
        if filename.endswith(".csv"):
            with open(out_path) as f:
                n = sum(1 for _ in f) - 1
        else:
            wb = openpyxl.load_workbook(out_path)
            n = wb.active.max_row - 1

        print(f"  ✓ {filename:<40}  {n:>3} transactions")

        # Copy to watch folder under the slug subfolder
        watch_slug_dir = WATCH_BASE / slug
        watch_slug_dir.mkdir(parents=True, exist_ok=True)
        dest = watch_slug_dir / filename
        dest.write_bytes(out_path.read_bytes())
        print(f"    → {dest}")

    print(f"""
Done! 6 files generated in:
  {OUT_DIR}

And copied to watch-folder subfolders:
  {WATCH_BASE}

Create these accounts in the app (http://localhost:3000/accounts) to trigger auto-import:

  Nickname              | Last Four | Expected slug
  ----------------------|-----------|-----------------------------
  Chase Checking        | 1234      | chase-checking-1234
  Chase Credit Card     | 5678      | chase-cc-5678
  BofA Checking         | 9012      | boa-checking-9012
  Amex Platinum         | 3456      | amex-platinum-3456
  Citi Card             | 7890      | citi-card-7890
  Generic Savings       | 2468      | generic-savings-2468
""")


if __name__ == "__main__":
    main()
